import os
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from procurement.models.procurementMethodDetail import ProcurementMethodDetail


class Command(BaseCommand):
    help = "Автоматический импорт пунктов закупок из одного источника напрямую из Excel-файла"

    def add_arguments(self, parser):
        # Добавляем обязательный аргумент пути к Excel-файлу
        parser.add_argument(
            'file_path',
            type=str,
            help='Абсолютный или относительный путь к файлу .xlsx с пунктами'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']

        # Проверяем физическое существование файла на диске
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Критическая ошибка: Файл по пути '{file_path}' не найден!"))
            return

        # Базовые верхнеуровневые методы (без parent)
        base_methods_names = [
            "Открытый конкурс",
            "Закрытый конкурс",
            "Электронный аукцион",
            "Процедура запроса ценовых предложений",
            "Закрытая процедура запроса ценовых предложений",
            "Биржевые торги",
            "Процедура закупки из одного источника",
        ]

        self.stdout.write(self.style.MIGRATE_LABEL("Запуск импорта процедур и парсинга Excel..."))

        try:
            # Открываем книгу Excel в режиме чтения данных (data_only=True игнорирует формулы)
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active  # Берем первый (активный) лист
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Не удалось прочитать Excel-файл. Ошибка: {str(e)}"))
            return

        # Выполняем атомарно в одной транзакции СУБД PostgreSQL
        with transaction.atomic():
            # 1. Загружаем или обновляем базовые методы верхнего уровня
            for method_name in base_methods_names:
                obj, created = ProcurementMethodDetail.objects.update_or_create(
                    name=method_name,
                    parent=None,
                    defaults={"is_active": True}
                )
                if created:
                    self.stdout.write(f"Создан базовый метод: {obj.name}")

            # 2. Получаем родительский метод для закупки из одного источника
            parent_method = ProcurementMethodDetail.objects.get(
                name="Процедура закупки из одного источника",
                parent=None
            )

            index = 1
            # 3. Построчно читаем колонку A (первый столбец)
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value

                # Пропускаем пустые ячейки
                if not cell_value:
                    continue

                # Приводим к строке и очищаем от мусорных переносов строк и пробелов по краям
                text_item = str(cell_value).strip()

                # Если в ячейке просто технический заголовок вроде "Наименование" или "Пункт", пропускаем его
                if text_item.lower() in ['наименование', 'пункт', 'текст']:
                    continue

                # Формируем название: обрезаем текст под лимит max_length=512 с запасом для селектов
                if len(text_item) > 150:
                    short_preview = text_item[:150] + "..."
                else:
                    short_preview = text_item

                display_name = f"Пункт {index}. {short_preview}"

                # Записываем в базу данных, защищаясь от дублирования при повторном импорте
                obj, created = ProcurementMethodDetail.objects.update_or_create(
                    name=display_name,
                    parent=parent_method,
                    defaults={"is_active": True}
                )
                if created:
                    self.stdout.write(f"Успешно добавлен из Excel: {obj.name}")
                    index += 1

        self.stdout.write(self.style.SUCCESS("Импорт из Excel-файла успешно завершен!"))
